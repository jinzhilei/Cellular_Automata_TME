'''
In the entire code, some key data can be recorded in a file. Specifically, what data to record can be modified or added according to requirements.
'''
import os
import numpy as np
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

class record:
    '''Record the cumulative quantities of various states of cells.'''
    def record(self,loop):
        contents = os.getcwd() + '/output_result/'
        filename = contents + 'number' + str(loop+1) + '.xls'
        result1 = len(self.normal_stem_cell_total)
        result2 = len(self.normal_cell_total)
        result3 = len(self.cancer_stem_cell_total)
        result4 = len(self.cancer_cell_total)
        with open(filename, 'a') as file_object:
            file_object.write(str(result1))
            file_object.write('\t')
            file_object.write(str(result2))
            file_object.write('\t')
            file_object.write(str(result3))
            file_object.write('\t')
            file_object.write(str(result4))
            file_object.write('\t')
            file_object.write('\n')

    '''Record the numbers of the specific cell growth processes for various states. The first to the fourth lines represent normal stem cells, normal cells, tumor stem cells, and tumor cells respectively. '''
    def record_result(self,loop):
        contents = os.getcwd() + '/output_result/'
        filename = contents + 'cell_result.xls'
        with open(filename, 'a') as file_object:
            for i in range(len(self.normal_stem_cell_total)):
                file_object.write(str(self.normal_stem_cell_total[i]))
                file_object.write('\t')
            file_object.write('\n')
            for i in range(len(self.normal_cell_total)):
                file_object.write(str(self.normal_cell_total[i]))
                file_object.write('\t')
            file_object.write('\n')
            for i in range(len(self.cancer_stem_cell_total)):
                file_object.write(str(self.cancer_stem_cell_total[i]))
                file_object.write('\t')
            file_object.write('\n')
            for i in range(len(self.cancer_cell_total)):
                file_object.write(str(self.cancer_cell_total[i]))
                file_object.write('\t')
            file_object.write('\n')

    '''Record the number of times normal cells have proliferated. '''
    def record_frequency1(self, loop):
        contents = os.getcwd() + '/output_result/'
        filename = contents + 'frequency1_' + str(loop + 1) + '.xls'
        with open(filename, 'a') as file_object:
            for i in range(len(self.frequency1)):
                file_object.write(str(self.frequency1[i]))
                file_object.write('\t')
                file_object.write('\n')

    '''Record the proliferation times of tumor cells.'''
    def record_frequency2(self, loop):
        contents = os.getcwd() + '/output_result/'
        filename = contents + 'frequency2_' + str(loop + 1) + '.xls'
        with open(filename, 'a') as file_object:
            for i in range(len(self.frequency2)):
                file_object.write(str(self.frequency2[i]))
                file_object.write('\t')
                file_object.write('\n')

    '''Record the mean, maximum, and minimum values of the microenvironment, viewed column - by - column.'''
    def record_m(self, loop):
        contents = os.getcwd() + '/output_result/'
        filename = contents + 'm' + str(loop + 1) + '.xls'
        m = np.mean(self.m)
        m1 = np.max(self.m)
        m2 = np.min(self.m)
        with open(filename, 'a') as file_object:
            file_object.write(str(m))
            file_object.write('\t')
            file_object.write(str(m1))
            file_object.write('\t')
            file_object.write(str(m2))
            file_object.write('\t')
            file_object.write('\n')

    '''Record all the values of the microenvironment at the final moment, with each row representing the microenvironment values for one day.'''
    def record_m_real(self, loop):
        contents = os.getcwd() + '/output_result/'
        filename = contents + 'm_real' + str(loop + 1) + '.xls'
        with open(filename, 'a') as file_object:
            for i in range(len(self.m)):
                file_object.write(str(self.m[i]))
                file_object.write('\t')
            file_object.write('\n')


    '''Record the numbers of tumor stem cells and tumor cells generated under different values of k11 and k12. '''
    def k_record(self, loop):
        contents = os.getcwd() + '/output_result/'
        filename = contents + 'k_num.xls'
        result1 = len(self.normal_stem_cell_total)
        result2 = len(self.normal_cell_total)
        result3 = len(self.cancer_stem_cell_total)
        result4 = len(self.cancer_cell_total)
        result = [result1, result2, result3, result4]
        with open(filename, 'a') as file_object:
            file_object.write(str(result))

    '''Record the quantities of various types of cells generated under different mutation rates and different microenvironments. '''
    def s_record(self, loop):
        contents = os.getcwd() + '/output_result/'
        filename = contents + 'm_t.xlsx'  # Note that the file extension should be.xlsx.

        if not os.path.exists(contents):
            os.makedirs(contents)

        # Create a new workbook or load an existing one.
        if os.path.exists(filename):
            workbook = load_workbook(filename)
        else:
            workbook = Workbook()

        # Get or create a worksheet.
        sheet_name = 'Sheet{}'.format(loop)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)

        result1 = len(self.normal_stem_cell_total)
        result2 = len(self.normal_cell_total)
        result3 = len(self.cancer_stem_cell_total)
        result4 = len(self.cancer_cell_total)

        # Append data to the worksheet.
        new_data = [result1, result2, result3, result4]
        sheet.append(new_data)

        # Delete the worksheet named "Sheet".
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']

        # Save the workbook to the specified path.
        workbook.save(filename)